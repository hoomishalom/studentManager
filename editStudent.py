import customtkinter as ctk
import openpyxl as xl
import os
from helperFunctions import createFileName, createDirName

class EditStudent(ctk.CTkFrame):
    def testValidations(self, e):
        widget = str(self.focus_get()).split(".!entry")[0]
        
        if (widget == str(self.studentNameEntry)): # checks if both widgets are the same
            self.validateName()
        # elif (widget == str(self.studentDateEntry)):
        #     self.validateDate()
        # elif (widget == str(self.studentHoursEntry) or widget == str(self.studentMinutesEntry)):
        #     self.validateTime()
        elif (widget == str(self.studentDefaultRateEntry)):
            self.validateRate()
            
            
    def validateName(self):
        self.students = os.listdir(self.studentsPath)
        
        if (self.studentNameEntry.get().lower() in [name.lower() for name in self.students]):
            self.studentName = self.studentNameEntry.get()
            self.defaultRate = self.getDefaultRate()
            self.studentDefaultRateEntry.delete(0, ctk.END)
            self.studentDefaultRateEntry.insert(0, str(self.defaultRate))
            self.validateRate()
            self.nameValid = True
            self.studentNameValidateCheckBox.select()
            self.getLessons()
        else:
            self.nameValid = False
            self.rateValid = False
            self.defaultRate = 0
            self.studentDefaultRateValidateCheckBox.deselect()
            self.studentNameValidateCheckBox.deselect()
            self.resetRate()
            self.getLessons()
            
            
            
    def validateRate(self):
        rate = self.studentDefaultRateEntry.get()
        
        if (rate == ""):
            rate = 0
        
        self.studentDefaultRateEntry.delete(0, ctk.END)
        self.studentDefaultRateEntry.insert(0, str(int(rate)))
        
        
        if (int(rate) > 0):
            self.rateValid = True
            self.studentDefaultRateValidateCheckBox.select()
        else:
            self.rateValid = False
            self.studentDefaultRateValidateCheckBox.deselect()
    
    
    def getDefaultRate(self):
        if self.studentName != "":
            wb = xl.load_workbook(os.path.join(createDirName(self.studentsPath, self.studentName), createFileName(self.studentName)))
            return wb.active[self.rateLocation].value
        else:
            return 0

    
    def getLessons(self):
        if self.nameValid:
            wb = xl.load_workbook(os.path.join(createDirName(self.studentsPath, self.studentName), createFileName(self.studentName)))
            temp = wb.active["A"]
            self.lessonList = tuple([temp[i].value for i in range(1, len(temp)) if temp[i].value != None])
            self.lessonPicker.configure(values = self.lessonList)
        else:
            self.lessonPicker.configure(values = [])
    
    
    def resetRate(self):
        self.studentDefaultRateEntry.delete(0, ctk.END)
        self.studentDefaultRateEntry.insert(0, str(self.defaultRate))
    
    
    def validateTimeInput(self, P):
        return (str.isdigit(P) or P == "")
    
    
    def submitDefaultRate(self):
        if self.rateValid:
            self.defaultRate = int(self.studentDefaultRateEntry.get())
            wb = xl.load_workbook(os.path.join(createDirName(self.studentsPath, self.studentName), createFileName(self.studentName)))
            wb.active["Q2"] = int(self.defaultRate)
            wb.save(os.path.join(createDirName(self.studentsPath, self.studentName), createFileName(self.studentName)))
    
    
    def __init__(self, master, studentsPath, rateLocation, **kwargs):
        super().__init__(master, **kwargs)
        
        self.master = master
        self.studentsPath = studentsPath
        self.rateLocation = rateLocation
        self.studentName = ""
        self.defaultRate = 0
        self.rateValid = False
        self.nameValid = False
        self.dateColumn = "A"
        self.hoursColumn = "B"
        self.minutesColumn = "C"
        self.rateColumn = "E"
        self.rateFrameHeight = kwargs.get("height") / 5
        self.rateFrameWidth = kwargs.get("width")
        self.lessonList = []
        
        
        self.frameNameLabelRate = ctk.CTkLabel(self, text="Edit Lesson", font=("Roboto", 24))
        self.frameNameLabelRate.pack(pady=10, padx=10, side="top")
        
        self.rateEditFrame = ctk.CTkFrame(master=self, height=self.rateFrameHeight, width=self.rateFrameWidth)
        self.rateEditFrame.propagate(False) 
        self.rateEditFrame.pack(pady=10, padx=10, side="top")
        
        self.studentNameLabel = ctk.CTkLabel(self.rateEditFrame, text="Name:")
        self.studentNameLabel.grid(row=0, column=0, pady=10, padx=10, sticky="W")
        self.studentNameEntry = ctk.CTkEntry(self.rateEditFrame)
        self.studentNameEntry.grid(row=0, column=1, pady=10, padx=10, sticky="E")
        self.studentNameValidateCheckBox = ctk.CTkCheckBox(self.rateEditFrame, text="", state="DISABLED", width = 30)
        self.studentNameValidateCheckBox.grid(row=0, column=2, padx=10)
        
        self.studentDefaultRateLabel = ctk.CTkLabel(self.rateEditFrame, text="Rate:")
        self.studentDefaultRateLabel.grid(row=1, column=0, padx=10, pady=10, sticky="W")
        self.studentDefaultRateEntry = ctk.CTkEntry(self.rateEditFrame, validate="all", validatecommand=(self.register(self.validateTimeInput), "%P"))
        self.studentDefaultRateEntry.grid(row=1, column=1, padx=10, pady=10, sticky="E")
        self.studentDefaultRateEntry.insert(0, str(self.defaultRate))
        self.studentDefaultRateValidateCheckBox = ctk.CTkCheckBox(self.rateEditFrame, text="", state="DISABLED", width = 30)
        self.studentDefaultRateValidateCheckBox.grid(row=1, column=2, padx=10)
        self.resetDefaultRateButton = ctk.CTkButton(self.rateEditFrame, text="Reset Rate", command=self.resetRate)
        self.resetDefaultRateButton.grid(row=1, column=3, padx=10, pady=10)
        self.rateSubmitButton = ctk.CTkButton(self.rateEditFrame, text="Submit Rate", command=self.submitDefaultRate)
        self.rateSubmitButton.grid(row=1, column=4, padx=10)
        
        self.studentLessonPickerLabel = ctk.CTkLabel(self.rateEditFrame, text="Lesson:")
        self.studentLessonPickerLabel.grid(row=2, column=0, padx=10, pady=10, sticky="W")
        self.lessonPicker = ctk.CTkComboBox(self.rateEditFrame, values=self.lessonList, state="readonly")
        self.lessonPicker.grid(row=2, column=1)
